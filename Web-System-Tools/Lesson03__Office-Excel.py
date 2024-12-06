from openpyxl import Workbook

def Save_Excel():
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "hello"
    sheet["B1"] = "world!"

    workbook.save(filename="data/myfile.xlsx")

def Read_Excel():
    from openpyxl import load_workbook
    workbook = load_workbook(filename="data/myfile.xlsx")
    print(workbook.sheetnames)
    # ['Sheet 1']

    sheet = workbook.active
    print(sheet)
    # < Worksheet; "Sheet 1" >

    print(sheet.title)


if __name__ == "__main__":
    Save_Excel()
    Read_Excel()
